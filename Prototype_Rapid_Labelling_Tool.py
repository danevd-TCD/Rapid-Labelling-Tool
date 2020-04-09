#---- documentation ---
# Pillow:       https://pillow.readthedocs.io/en/stable/index.html
# OpenPyXL:     https://openpyxl.readthedocs.io/en/stable/index.html
# PySimpleGUI:  https://pysimplegui.readthedocs.io/en/latest/
# excel2json-3: https://pypi.org/project/excel2json-3/

import PySimpleGUI as sg #PySimpleGUI is framework for entire GUI element of this application. 
import glob #recursive file/folder search functionality. native to Python
import csv #deliminate csv output of glob using pythons' native csv support
from pathlib import Path
import os
from PIL import Image #PILlow image library for image-based operations
import shutil
import itertools
from openpyxl import Workbook #library for saving excel workbooks
from excel2json import convert_from_file #for converting .xlsx to .json

#define small/med/large image sizes, for UI customisation
sizeSmall = 400
sizeMed = 500
sizeLarge = 600

#colorblind options
#normal vision
defaultCol_ButtonProceed = ('black','#a1d784')
defaultCol_ButtonExit = ('white','#aa2d2a')
defaultCol_textBG = '#f7e0a5'

#Deuteranopia (red-green)
Deuteranopia_ButtonProceed= ('black','#f0e442')
Deuteranopia_ButtonExit = ('white','#0072b2')
Deuteranopia_textBG = '#bae8ff'

#Protanopia (red-green)
Protanopia_ButtonProceed= ('black','#f0e442')
Protanopia_ButtonExit = ('white','#0072b2')
Protanopia_textBG = '#bae8ff'

#tritanopia (blue-yellow)
Tritanopia_ButtonProceed = ('black','#f0e442')
Tritanopia_ButtonExit = ('white','#0072b2')

ButtonProceed=defaultCol_ButtonProceed #ColVal is current colorscheme for buttons
ButtonExit=defaultCol_ButtonExit
textBG = defaultCol_textBG

VarUISize = sizeSmall #define standard width/height by relating to sizeSmall

# theme of the GUI. themes: https://user-images.githubusercontent.com/46163555/70382042-796da500-1923-11ea-8432-80d08cd5f503.jpg
sg.theme('LightGrey1')

StyleColour = '#fafafa'
sg.theme_element_background_color(StyleColour)
sg.theme_text_element_background_color(StyleColour)
sg.theme_background_color(StyleColour)

sg.SetOptions(element_padding=(4, 4)) #UI element padding - global

###LauncherWindow###
l_tab1 = [
            [sg.Text('Please select an application mode:')],
            [sg.Text('')],
            [sg.Text('Multi-label classification (MLC)', size=(30,1)), sg.Button('Launch MLC', key='DL_Launch')],
            [sg.Text('Object location classification (OLC)', size=(30,1)), sg.Button('Launch OLC', key='Draw_Launch')],
            ]

l_tab2 = [
            [sg.Text('Output file format:')],
            [sg.OptionMenu(('.xlsx','.json','.xlsx + .json'),key="OutputFormat", auto_size_text=None, size=(40,1), default_value='.xlsx + .json')],
            [sg.Text('Colourblind setting:')],
            [sg.OptionMenu(('None','Deuteranopia','Protanopia','Tritanopia'),key="ColourBlindSettingLaunch", auto_size_text=None, size=(40,1))],
            [sg.Text('Image size in interfaces:')],
            [sg.OptionMenu(('Small','Medium','Large'),key="SizeImageUI", auto_size_text=None, size=(40,1))]
         ]

LauncherLayout = [
                    [sg.TabGroup([[sg.Tab('Main',l_tab1), sg.Tab('Options',l_tab2)]])]
                 ]

LaunchWindow = sg.Window('Rapid Labelling Prototype', LauncherLayout, font=("Tahoma",12), finalize=True)

while True:
    LaunchEv, LaunchVals = LaunchWindow.Read(timeout=100)

    #colourblind setting
    if LaunchVals['ColourBlindSettingLaunch'] == 'None':
        ButtonProceed=defaultCol_ButtonProceed 
        ButtonExit=defaultCol_ButtonExit
        textBG=defaultCol_textBG

    elif LaunchVals['ColourBlindSettingLaunch'] == 'Deuteranopia':
        ButtonProceed=Deuteranopia_ButtonProceed
        ButtonExit=Deuteranopia_ButtonExit
        textBG=Deuteranopia_textBG

    elif LaunchVals['ColourBlindSettingLaunch'] == 'Protanopia':
        ButtonProceed=Protanopia_ButtonProceed
        ButtonExit=Protanopia_ButtonExit
        textBG=Protanopia_textBG

    elif LaunchVals['ColourBlindSettingLaunch'] == 'Tritanopia':
        ButtonProceed=Tritanopia_ButtonProceed
        ButtonExit=Tritanopia_ButtonExit

    if LaunchEv == 'LaunchExit': #if user hits 'Exit' button in first window, close window1 and break loop, ending program
        LaunchWindow.close()
        break

    if LaunchVals['SizeImageUI'] == 'Small': #if value from 'SizeImageUI' in tab2_layout = small,
        VarUISize = sizeSmall #set VarUISize = sizeSmall.
    elif LaunchVals['SizeImageUI'] == 'Medium':
        VarUISize = sizeMed
    elif LaunchVals['SizeImageUI'] == 'Large':
        VarUISize = sizeLarge


    if LaunchEv == 'DL_Launch': #if main datalabeller is launched:
        Tagbook1 = Workbook()
        MainSheet = Tagbook1.active
        MainSheet.title="Data"
        MainSheet.cell(row=1, column=1, value=('Filename'))

        LaunchWindow.hide()
        #individual labeller layout:
        tab1_layout = [
                        [sg.Text('Please enter the number of different data annotation categories, and the number of subsequent definitions per category, that you wish to define for this session.',size=(55,3))],
                        [sg.Text('Data annotation categories:',size=(23,1),pad=((1,1),(20,5))),sg.OptionMenu((range(1,15)),pad=((1,1),(20,5)), key="InputNum")],
                        [sg.Text('Definitions per category:',size=(23,1),pad=((1,1),(5,20))),sg.OptionMenu((range(2,15)),pad=((1,1),(5,20)), key="RowNum")],
                        [sg.Text('Please click "Define Annotations" to proceed',size=(36,1)),sg.Button('Define Annotations',button_color=ButtonProceed,key="mainProceed")],
                     ]

        #combine tabs into output layout, which sg.Window command then draws as a GUI:
        layout = [  
                    [sg.Text('Please press the "Next" button to continue',size=(50,1),justification='center', visible=False, key="NextText")],
                    [sg.Button('Exit',key="MainExit",button_color=ButtonExit,pad=((3,420),(3,3)))]
                 ]

        # Create the Window
        window1 = sg.Window('MLC - Setup window', tab1_layout + layout, font=("Tahoma", 12),  finalize=True)

        #This logic loop is significant: it dicates the multi-window operation of the application.
        while True:
            ev1, vals1 = window1.Read(timeout=100) #read all events ("ev1") and values ("vals1") of window1, every 100 miliseconds.

            if ev1 == 'mainProceed':
                headingsArray=[]
                TableDataArray = []
                if len(vals1['InputNum']) > 0:
                    window1.Hide()

                    input_rows = [ 
                                   #We define a H1-R1 notation style; that is, acess the value of Header 1, Row 2, we can use valsDef['H1-R2']. For Header 3, row 7, it would be: valsDef['H3-R7']
                                   #To get the value of just the header, which is to say the header name, we use H1-R1, as the first row is composed of the headers themselves. 
                                   [sg.Text('Row ' + str((row + 1)), size=(12,1))]+[sg.Input(size=(14,1) , key='H'+ str(col) + '-' + 'R' + str((row+1))) for col in range(1,(int(vals1['InputNum']) + 1))] for row in range(1,(int(vals1['RowNum']) + 1))
                                 ]

                    for i in range(1,(int(vals1['InputNum']) + 1)):
                        headingsArray.append(i) #append numerical value for current header based on int value of users' input for number of label definitions
                
                    blurb =   [[sg.Text('Welcome to the data annotation definition area. Please specify the category, e.g "Age of person", and add appropriate definitions in the rows below, e.g "young", "teenager" , "old" etc.', size=(80,None))]]

                    header =  [
                                [sg.Text('Category:', size=(12,1))] + [sg.Input('Category ' + str(h), size=(14,1), text_color='#003c71', background_color=textBG, key='H' + str(h) + '-R1') for h in headingsArray]  # build header layout
                              ] 

                    Proc = [
                             [sg.Text('When you are happy with the data entered above, please click on the "Proceed" button')],
                             [sg.Button('Back',button_color=ButtonExit),sg.Button('Proceed',button_color=ButtonProceed)]
                           ]

                    testLayout = blurb + header + input_rows + Proc
           
                    windowDefine = sg.Window('MLC - Annotation definition',testLayout, element_justification='center', font=("Tahoma", 12))
                    while True:
                        evDef, valsDef = windowDefine.Read(timeout=100)

                        if evDef =='Proceed':
                            #nested iteration using itertools, to achieve [0,0 , 0,1, 0,2, 1,0, 1,1 , 1,2, 2,0, 2,1] etc.
                            for i, j in itertools.product(range(1,int(vals1['InputNum'])+1), range(1,int(vals1['RowNum'])+2) ):
                                currentVal = str(valsDef['H'+str(i)+'-'+'R'+str(j)])
                                TableDataArray.append(currentVal)

                            windowDefine.Hide()

                            #Interrim window is the windwo that loads between user selecting to launch the main UI, and the main UI opening.
                            #the interrim window is used to prompt the user to select a (valid) folder of source images for the next UI stage
                            layoutInterrim = [ 
                                                [sg.Text('Please select the source folder of the image data to be annotated. Click on the "Browse" button and navigate to the appropriate folder', size=(55,3))],
                                                [sg.Input('',key='FolderLoc'), sg.FolderBrowse()], 
                                                [sg.Text('When you have entered a valid destination, please press the green "Continue" button',size=(55,2))],
                                                [sg.Button('Back',button_color=ButtonExit,key="InterrimBack"),sg.Button('Continue',disabled=True,pad=((360,3),(3,3)), key="HiddenContinue")]
                                             ]

                            #we tell pySimpleGUI to draw windowInterrim, using layoutInterrim
                            windowInterrim = sg.Window('MLC - Select source', layoutInterrim, element_justification='center',font=("Tahoma", 12),finalize=True)

                            while True:
                                evInt, valsInt = windowInterrim.Read(timeout=100) #Read all events and inputs ever 100 milliseconds

                                if evInt is None or evInt == 'InterrimBack': #if user hits "back" during folder source location prompt, go back to table array definition and reset TableDataArray to prevent duplication
                                    windowInterrim.Close()
                                    TableDataArray = []
                                    windowDefine.UnHide()
                                    break

                                def resetOptionMenu(): #this function will reset optionmenus in between image forward/backward navigation
                                    for i in range(len(TableDataArray)):
                                       if i % (int(vals1['RowNum'])+1) ==0:
                                          window2[TableDataArray[i]].update(values=(selectArray+TableDataArray[i+1:i+(1+int(vals1['RowNum']))]))

                                #show hidden continue button once user loads any directory with length > 1
                                if len(valsInt['FolderLoc']) > 1:
                                    windowInterrim['HiddenContinue'].update(disabled=False,button_color=ButtonProceed)

                                if evInt == 'HiddenContinue':
                                    InitialSet = 0
                                    windowInterrim.Close()
                                    img_count = 0
                                    row_count = img_count + 2

                                    pathInput = (valsInt['FolderLoc'])
                                    makeDirLoc = os.path.normpath(pathInput +'/temp') #keep an eye on this: will it work in other OS due to backslash/forwardslash disparities?
                                    pathSnip = len(makeDirLoc) #a variable used later, to strip folder location from filename

                                    #this logic loop checks to see if a /temp folder already exists at target destination
                                    #if so, delete it and its contents; if not, pass onto next loop
                                    try:
                                        shutil.rmtree(makeDirLoc)
                                    except OSError:
                                        pass

                                    #this loop attempts to make a folder
                                    try:
                                       os.makedirs(makeDirLoc)
                                    except FileExistsError:
                                        print('folder already exists')

                                    files = [f for f in glob.glob(pathInput + "**/*.jpg", recursive=True)]
                                    csv_reader = csv.reader(files, delimiter=',')
                                    line_count = 0
                                    ImgArray =[]

                                    #read each row of delimited files with .jpg filetype
                                    for row in csv_reader:
                                       PathLoc = os.path.normpath(f'{row[0]}')
                                       EditPathLoc = PathLoc + ".png" 
                                       ImgArray.append(PathLoc)
                                       line_count += 1

                                       try:

                                           im = Image.open(PathLoc) #.resize(VarUISize).save(EditPathLoc, "PNG") #open each image; save to root folder #####RESIZE####
                                           y_scale = VarUISize / im.height
                                           x_scale = VarUISize / im.width

                                           if y_scale < x_scale: #all portrait photos: where height is larger than width
                                               (width,height) = (int(im.width * y_scale), int(im.height * y_scale))
                                               im.resize((width,height)).save(EditPathLoc, "PNG")

                                           elif y_scale > x_scale: #all landscape photos
                                               (width,height) = (int(im.width * x_scale), int(im.height * x_scale))
                                               im.resize((width,height)).save(EditPathLoc, "PNG")

                                           elif y_scale == x_scale: #for photos that happen to be perfectly square
                                               (width,height) = (int(im.width * y_scale), int(im.height * y_scale))
                                               im.resize((width,height)).save(EditPathLoc, "PNG")

                                           shutil.move(EditPathLoc,makeDirLoc) #move the newly-generated .pngs to /temp folder

                                       except IOError:
                                           print("IOError at Image rewrite functionality")
                                           pass

                                    for i, j in itertools.product(range(1,int(vals1['InputNum'])+1), range(1,int(vals1['RowNum'])+2) ):
                                      currentVal = str(valsDef['H'+str(i)+'-'+'R'+str(j)])        

                                    selectArray = ['--select--'] #insert a piece of text at the top of the OptionMenu to remind/show user that this field has not been changed 

                                    genCol =[
                                            [sg.Text(TableDataArray[i])] + [sg.OptionMenu(selectArray+TableDataArray[i+1:i+(1+int(vals1['RowNum']))],key=TableDataArray[i])] for i in range(len(TableDataArray)) if i % (int(vals1['RowNum'])+1) ==0
                                            ]

                                    print(TableDataArray)

                                    frameButtons = [
                                                     [sg.Button('Previous image',key="PrevImg", pad=((1,225),(3,3))),sg.Button('Next image',key="NextImg", pad=((225,1),(3,3)))],
                                                     [sg.Text('Image ' + str(img_count + 1) + ' of ' + str(len(ImgArray)),key="CounterKey")],
                                                     [sg.Button('Save',key="WriteToExcel",button_color=ButtonProceed)]
                                                   ]

                                    layoutUI = [
                                                [sg.Image('', key='ImageUpdate'),
                                                 sg.Column(genCol,size=(300,400),scrollable=True,vertical_scroll_only=True, element_justification='right')],
                                                [sg.Column(layout=frameButtons, element_justification='center')]
                                               ]

                                    window2 = sg.Window('MLC - Rapid Data Labelling Interface', layoutUI, font=("Tahoma", 12),finalize=True)        #we tell pySimpleGUI to draw window2, using layout2

                                    #Generate headers on first row of excel sheet:
                                    TitleArray = TableDataArray[::(int(vals1['RowNum'])+1)] #gather every nth element, where n=number of rows 
                                    for i in range(0,(int(vals1['InputNum']))): 
                                        MainSheet.cell(row=1, column=(i+2), value=TitleArray[i])

                                    while True:
                                        ev2, vals2 = window2.Read(timeout=100) #as in window1, read all events and values input by the user into window 2
                                        dest_filename = os.path.normpath(makeDirLoc + '//MLC_Output.xlsx')
                                        line_count = 0
                                        pathInput = makeDirLoc
                                        files = [f for f in glob.glob(pathInput + "**/*.png", recursive=True)]
                                        csv_reader = csv.reader(files, delimiter=',')
                                        ImgArray =[]
                                        for row in csv_reader:
                                            ImgArray.append(os.path.normpath(f'{row[0]}'))
                                            line_count += 1

                                        if ev2 is None or ev2 == 'Exit': #if user hits exit, close window 2, unhide window1
                                            window2.Close()
                                            window1.UnHide()
                                            break

                                        if InitialSet == 0:
                                            window2['ImageUpdate'].update(f'{ImgArray[0]}')
                                            InitialSet = 1
                                        else:
                                            pass

                                        window2['CounterKey'].update('Image ' + str(img_count + 1) + ' of ' + str(len(ImgArray)))

                                        #the following logic loop disables the "previous image" button if the user is on the first image
                                        if img_count == 0:
                                            window2['PrevImg'].update(disabled=True)
                                        elif img_count > 0:
                                            window2['PrevImg'].update(disabled=False)
                                
                                        #the following logic loop disables the "next image" button if the user is on the last image
                                        if img_count == (len(ImgArray)-1):
                                            window2['NextImg'].update(disabled=True)
                                        else:
                                            window2['NextImg'].update(disabled=False)

                                        if ev2 == 'PrevImg':

                                             resetOptionMenu() #we invoke this function to cause the optionmenus to reset upon navigating to a previous image

                                            #logic loop to prevent negative indexing in python
                                             if img_count == 1: #if at second-last image (0, [1], 2, 3..), we are going back to 0 as this loop...
                                                                #... is inside the 'Previous image' logic loop button press incident
                                                img_count = 0   #set img_count to 0, then continue to next
                                                row_count = img_count + 2

                                             elif img_count == 0: #if already at 0 img count (e.g on first launch), set to 0 and prompt user
                                                img_count = 0
                                                row_count = img_count + 2

                                             else: #for n>1, decrement as usual
                                                img_count += -1
                                                row_count = img_count + 2

                                             #print(img_count)
                                             print(row_count)

                                             try: #try to update window...
                                                window2['ImageUpdate'].update(f'{ImgArray[img_count]}')
                                             except IndexError: #..unless we get an indexerror, aka have reached the first image
                                                 sg.popup('Reached first image')
                                                 pass

                                        if ev2 == 'NextImg':

                                             resetOptionMenu() #we invoke this function to cause the optionmenus to reset upon navigating to a previous image

                                             try: #try to update window...
                                                img_count += 1
                                                row_count = img_count + 2
                                                print(str(img_count) + ': incremented Try state') #diagnostic: print if successul increment
                                                window2['ImageUpdate'].update(f'{ImgArray[img_count]}')
                
                                             except IndexError: #..unless we get an indexerror, aka have reached the last image
                                                 img_count=len(ImgArray)-1
                                                 row_count = img_count + 2
                                                 print(str(img_count) + ': IndexError reduction') #diagnostic: print if reduced val through IndexError exception
                                                 pass

                                        if ev2 =='WriteToExcel':
                                            if LaunchVals['OutputFormat'] == '.xlsx':
                                                Tagbook1.save(filename=dest_filename)
                                            elif LaunchVals['OutputFormat'] == '.json':
                                                Tagbook1.save(filename=dest_filename)
                                                convert_from_file(dest_filename)
                                                os.remove(dest_filename)
                                            elif LaunchVals['OutputFormat'] == '.xlsx + .json':
                                                Tagbook1.save(filename=dest_filename)
                                                convert_from_file(dest_filename)

                                        #n.b likely source of issues here with different OS's and their different file structures. os.path.normpath may be required
                                        writeVal = [sub[ pathSnip+1:-4] for sub in ImgArray] #remove last 4 chars aka ".png", and first n characters where pathSnip is length of folder location/
                                        
                                        MainSheet.cell(column=1, row=row_count,  value= writeVal[img_count])
                                        #for i between 0 and last header, select the (ith + 2) column {as indexing starts at 0} and write to it the current value of the
                                        # OptionMenu with key equal to (i * number of rows), as optionmenus are defined with a key equal to the header they are assigned next to
                                        for i in range(0,(int(vals1['InputNum']))):
                                            MainSheet.cell(row=row_count, column=(i+2), value=str(vals2[TableDataArray[i*(1+len(range(int(vals1['RowNum']))))]] ))

                        if evDef is None or evDef =='Back':
                            windowDefine.Close()
                            window1.UnHide()
                            break #nb make sure to include "break" at end of exit/back button loop, else entire application will hang

                else: #popup alert telling user to enter more than 1 character
                    sg.Popup('Error: Enter at least 1 number', keep_on_top=True)
        
        
            if ev1 is None: #failsafe logic loop element, do not touch
                break

            if ev1 == 'MainExit': #go back to initial launch screen
                window1.close()
                LaunchWindow.UnHide()
                break


    if LaunchEv == 'Draw_Launch': #if rectangle drawer is launched
       Tagbook2 = Workbook()
       MainSheet = Tagbook2.active
       MainSheet.title="Draw_Data"
       MainSheet.cell(row=1, column=1, value=('Filename'))

       LaunchWindow.hide()
       
       draw_tab1_layout = [
                       [sg.Text('Please enter the number of different object categorisation elements',size=(40,2))],
                       [sg.Text('Categories:',size=(15,1),pad=((1,1),(20,5))),sg.OptionMenu((range(1,15)),pad=((1,1),(20,5)), key="DrawNum"),sg.Button('Go',button_color=ButtonProceed,key="DrawGo",pad=((1,1),(20,5)))] #DrawNum = number of rectangle definitions
                    ]
            
       layout = [  
                   [sg.Text('Please press the "Next" button to continue',size=(50,1),justification='center', visible=False, key="NextTextDraw")],
                   [sg.Button('Exit',key="DrawExit",button_color=ButtonExit,pad=((3,420),(3,3)))]
                ]
       
       # Create the Window
       DrawLauncher = sg.Window('OLC - Setup window', draw_tab1_layout + layout, font=("Tahoma", 12),  finalize=True, element_justification='left')
       
       while True:
           DrawLaunchEvent, DrawLaunchVals = DrawLauncher.Read(timeout=100) 
       
           if DrawLaunchEvent == 'DrawGo':
               DrawLauncher.hide()

               DrawTableDataArray = []
      
               draw_input_rows = [ 
                              #values: key is "R1" , "R2" etc, DrawWindDefVals["R2"]
                              [sg.Text('Category ' + str((row)), size=(12,1))]+[sg.Input(size=(14,1) , key='R' + str((row)))]for row in range(1,(int(DrawLaunchVals['DrawNum']) + 1))
                            ]
      
               draw_blurb =   [[sg.Text('Welcome to the object category definition area. Please specify the category, e.g "Age of person"', size=(80,None))]]
      
               draw_Proc = [
                        [sg.Text('When you are happy with the data entered above, please click on the "Proceed" button')],
                        [sg.Button('Back',button_color=ButtonExit),sg.Button('Proceed',button_color=ButtonProceed)]
                      ]
      
               draw_layout = draw_blurb + draw_input_rows + draw_Proc
               draw_windowDefine = sg.Window('OLC - Category definition',draw_layout, element_justification='center', font=("Tahoma", 12))
               
               while True:
                   DrawWindDefEvent, DrawWindDefVals = draw_windowDefine.Read(timeout=100)

                   if DrawWindDefEvent == 'Proceed':
                    for j in range(1,int(DrawLaunchVals['DrawNum'])+1):
                        DrawcurrentVal = str(DrawWindDefVals['R'+str(j)]) #e.g :DrawWindDefVals[R4] is the fourth user-entered value in draw_input_rows
                        DrawTableDataArray.append(DrawcurrentVal) #then append these values to DrawTableDataArray
                    
                    print(DrawTableDataArray)
                    draw_windowDefine.Hide()
                    
                    #layout for folder loc
                    DrawFolderLoc = [ 
                                        [sg.Text('Please select the source folder of the image data to be categorised. Click on the "Browse" button and navigate to the appropriate folder', size=(55,3))],
                                        [sg.Input('',key='FolderLoc'), sg.FolderBrowse()], 
                                        [sg.Text('When you have entered a valid destination, please press the green "Continue" button',size=(55,2))],
                                        [sg.Button('Back',button_color=ButtonExit,key="InterrimBack"),sg.Button('Continue',disabled=True,pad=((360,3),(3,3)), key="HiddenContinue")]
                                     ]
                    
                    DrawFolderLoc= sg.Window('OLC - Select source', DrawFolderLoc, element_justification='center',font=("Tahoma", 12),finalize=True)
                    
                    #Generate headers on first row of excel sheet:
                    for i in range(0,(int(DrawLaunchVals['DrawNum']))): 
                        MainSheet.cell(row=1, column=(i+2), value=DrawTableDataArray[i])

                    while True:
                        DrawLocEv, DrawLocVals= DrawFolderLoc.Read(timeout=100) 

                        if DrawLocEv is None or DrawLocEv == 'InterrimBack': #if user hits "back" during folder source location prompt, go back to draw def window and reset DrawTableDataArray
                            DrawFolderLoc.Close()
                            DrawTableDataArray = []
                            draw_windowDefine.UnHide()
                            break

                        #enable continue button once user loads any directory with length > 1
                        if len(DrawLocVals['FolderLoc']) > 1:
                            DrawFolderLoc['HiddenContinue'].update(disabled=False,button_color=ButtonProceed)

                        if DrawLocEv == 'HiddenContinue':
                            img_count = 0
                            row_count = img_count + 2

                            ImgWidth = [] #output image width
                            ImgHeight =[] #output image height
                            ScaleFactor =[] #the y_scale or x_scale value used to resize image

                            ClickCount = 1  #counter for num of clicks
                            InitialSet = 0
                            DrawFolderLoc.Close()
                            img_count = 0
                            
                            pathInput = (DrawLocVals['FolderLoc'])
                            makeDirLoc = os.path.normpath(pathInput +'/temp') #keep an eye on this: will it work in other OS due to backslash/forwardslash disparities?
                            pathSnip = len(makeDirLoc) #a variable used later, to strip folder location from filename
                            
                            #this logic loop checks to see if a /temp folder already exists at target destination
                            #if so, delete it and its contents; if not, pass onto next loop
                            try:
                                shutil.rmtree(makeDirLoc)
                            except OSError:
                                pass
                            
                            #this loop attempts to make a folder
                            try:
                               os.makedirs(makeDirLoc)
                            except FileExistsError:
                                print('folder already exists')
                            #except IOError:
                            #    print('Access denied')
                            
                            files = [f for f in glob.glob(pathInput + "**/*.jpg", recursive=True)]
                            csv_reader = csv.reader(files, delimiter=',')
                            line_count = 0
                            ImgArray =[]                            

                            #read each row of delimited files with .jpg filetype
                            for row in csv_reader:
                               PathLoc = os.path.normpath(f'{row[0]}')
                               EditPathLoc = PathLoc + ".png" 
                               ImgArray.append(PathLoc)
                               line_count += 1
                            
                               try:
                                   im = Image.open(PathLoc) 
                                   y_scale = VarUISize / im.height
                                   x_scale = VarUISize / im.width
                            
                                   if y_scale < x_scale: #all portrait photos: where height is larger than width
                                       (width,height) = (int(im.width * y_scale), int(im.height * y_scale))
                                       ScaleFactor.append(y_scale)
                                       ImgWidth.append(width)
                                       ImgHeight.append(height)
                                       im.resize((width,height)).save(EditPathLoc, "PNG")
                            
                                   elif y_scale > x_scale: #all landscape photos
                                       (width,height) = (int(im.width * x_scale), int(im.height * x_scale))
                                       ScaleFactor.append(x_scale)
                                       ImgWidth.append(width)
                                       ImgHeight.append(height)
                                       im.resize((width,height)).save(EditPathLoc, "PNG")
                            
                                   elif y_scale == x_scale: #for photos that happen to be perfectly square
                                       (width,height) = (int(im.width * y_scale), int(im.height * y_scale))
                                       ScaleFactor.append(y_scale)
                                       ImgWidth.append(width)
                                       ImgHeight.append(height)
                                       im.resize((width,height)).save(EditPathLoc, "PNG")
                            
                                   shutil.move(EditPathLoc,makeDirLoc) #move the newly-generated .pngs to /temp folder
                           
                               except IOError:
                                   print("IOError at Image rewrite functionality")
                                   pass
                            
                            for j in range(1,int(DrawLaunchVals['DrawNum'])+1):
                              currentVal = str(DrawWindDefVals['R'+str(j)])  
                            
                            welcomeText = [
                                          [sg.Text('Welcome to the OLC Data Categorisation UI. Please select an object category from the buttons on the right. To define an object location, click the top left and bottom right points of a rectangle bounding the object inside the image. To reset, press "Reset"',size=(80,4))],  
                                          ]

                            #these are the box selection drawing buttons made up using DrawTableDataArray
                            genButton =[
                                         [sg.Button(DrawTableDataArray[i],key=DrawTableDataArray[i])] for i in range(len(DrawTableDataArray))# if i % (int(vals1['RowNum'])+1) ==0
                                       ]

                            #these are the nav buttons
                            frameButtons = [
                                             [sg.Button('Previous image',key="PrevImg", pad=((1,225),(3,3))),sg.Button('Next image',key="NextImg", pad=((225,1),(3,3)))],
                                             [sg.Text('Image ' + str(img_count + 1) + ' of ' + str(len(ImgArray)),key="CounterKey")],
                                             [sg.Button('Save',key="WriteToExcel",button_color=ButtonProceed),sg.Button('Reset', key="Reset2")]
                                           ]

                            DrawlayoutUI = [
                                             [sg.Graph((VarUISize,VarUISize), (0, VarUISize), (VarUISize, 0), key='-GRAPH-', enable_events=True, drag_submits=False, float_values=True),
                                              sg.Column(genButton,size=(150,VarUISize),scrollable=True,vertical_scroll_only=True, element_justification='left')],
                                             [sg.Column(layout=frameButtons, element_justification='center')]
                                           ]

                            def GraphFunc(): #function to call reset of window
                                graph.DrawImage(filename=f'{ImgArray[img_count]}', location=(0,0))
                            
                            MainDrawWindow = sg.Window('OLC - Data categorisation UI', DrawlayoutUI, font=("Tahoma", 12),finalize=True)       #create MainDrawWindow

                            g = MainDrawWindow['-GRAPH-'] #shorthand annotation variable, used later
                            graph = MainDrawWindow.Element("-GRAPH-") #similiar to above, but performs slightly different operation
                            MainDrawWindow.Finalize() #.Finalize() is necessary to allow graph element alteration 
                            PosListArray = [[] for i in range(len(DrawTableDataArray))] #Position List Array: we generate n empty lists, where n = number of buttons
                            DrawTextVal = DrawTableDataArray[0] #initial DrawTextVal is the first (zeroth) item in DrawTableDataArray
                            button_val = 0 #initial button number is the first (zeroth) button

                            while True: #loop to launch MainDrawWindow
                                MainDrawEvs, MainDrawVal = MainDrawWindow.read(timeout=85)

                                ##successfully getting button values and i-values per button hit, based on button order down i-list
                                for i in range (0,len(DrawTableDataArray)):
                                    if MainDrawEvs == DrawTableDataArray[i]:
                                        DrawTextVal = DrawTableDataArray[i] #DrawTextVal is the currently-selected-button's text value in DrawTableDataArray
                                        button_val = int(i)
                                        MainDrawWindow[DrawTableDataArray[button_val]].update(button_color=('#2b2b2b','#f5f5f5')) #active button is white background with black text
                                        
                                        #the following logic loop sets all non-active buttons to the default background colour and font colour
                                        #for all buttons with button_val *below* current active button
                                        for j in range (0,button_val):
                                            MainDrawWindow[DrawTableDataArray[j]].update(button_color=('#ffffff','#0079d3')) 
                                        #for all buttons with button_val *above* current active button
                                        for h in range (button_val+1,len(DrawTableDataArray)):
                                            MainDrawWindow[DrawTableDataArray[h]].update(button_color=('#ffffff','#0079d3'))
                                        pass
                                
                                mouse = MainDrawVal['-GRAPH-']
                                #reset event
                                if MainDrawEvs == 'Reset2':
                                    g.Erase()
                                    GraphFunc()
                                    PosListArray = [[] for i in range(len(DrawTableDataArray))]
                                
                                #main mouse event
                                if MainDrawEvs == '-GRAPH-' and mouse[1] < ImgHeight[img_count] and mouse[0] < ImgWidth[img_count]: #the "and" statements confine the loop to only work if within the image
                                    if mouse == (None, None):
                                        continue
                                    if ClickCount % 2 > 0: #odd
                                        top_left = (mouse[0], mouse[1])
                                    else:
                                        bottom_right = (mouse[0], mouse[1])
                                        g.draw_rectangle((top_left),(bottom_right), line_color='#ccff00', line_width=3) #ccff00 is fluorescent yellow/lime green, stands out against almost any colour
                                        g.draw_text(DrawTextVal,
                                        ((top_left[0]),(top_left[1]-10)), font='Courier 12', color='#ccff00')
                                        top_left_scaled = (int(top_left[0]*(1/ScaleFactor[img_count])), int(top_left[1]*(1/ScaleFactor[img_count])))
                                        bottom_right_scaled = (int(bottom_right[0]*(1/ScaleFactor[img_count])), int(bottom_right[1]*(1/ScaleFactor[img_count])))
                                        PosListArray[button_val].append((top_left_scaled,bottom_right_scaled)) #Inside of a (), we append (top_left,bottom_right) of each rectangle to PosListArray

                                    ClickCount += 1
                                    print(ClickCount)

                                dest_filename = os.path.normpath(makeDirLoc + '//OLC_Output.xlsx')
                                line_count = 0
                                pathInput = makeDirLoc
                                files = [f for f in glob.glob(pathInput + "**/*.png", recursive=True)]
                                csv_reader = csv.reader(files, delimiter=',')
                                ImgArray =[]
                                for row in csv_reader:
                                    ImgArray.append(os.path.normpath(f'{row[0]}'))
                                    line_count += 1

                                if MainDrawEvs is None or MainDrawEvs == 'Exit': 
                                    window2.Close()
                                    window1.UnHide()
                                    break
                                
                                #upon first launch, run this to set to 0th image
                                if InitialSet == 0:
                                    graph.DrawImage(filename=f'{ImgArray[0]}', location=(0,0))#draw the first image using the 0th element of ImgArray
                                    InitialSet = 1
                                else:
                                    pass

                                #update img counter
                                MainDrawWindow['CounterKey'].update('Image ' + str(img_count + 1) + ' of ' + str(len(ImgArray)))
                                
                                #the following logic loop disables the "previous image" button if the user is on the first image
                                if img_count == 0:
                                    MainDrawWindow['PrevImg'].update(disabled=True)
                                elif img_count > 0:
                                    MainDrawWindow['PrevImg'].update(disabled=False)
                                
                                #the following logic loop disables the "next image" button if the user is on the last image
                                if img_count == (len(ImgArray)-1):
                                    MainDrawWindow['NextImg'].update(disabled=True)
                                else:
                                    MainDrawWindow['NextImg'].update(disabled=False)
                                
                                if MainDrawEvs == 'PrevImg':
                                     
                                    #reset ClickCount if primed (i.e user has clicked for top-left rectangle pos) when navigating images.
                                    #this prevents drawing bounding box outside image area in edge cases
                                     if ClickCount % 2 == 0:
                                        ClickCount += 1

                                     PosListArray = [[] for i in range(len(DrawTableDataArray))]
                                     if img_count == 1: #if at second-last image (0, [1], 2, 3..), we are going back to 0 as this loop...
                                                        #... is inside the 'Previous image' logic loop button press incident
                                        img_count = 0   #set img_count to 0, then continue to next
                                        row_count = img_count + 2
                                
                                     elif img_count == 0: #if already at 0 img count (e.g on first launch), set to 0 and prompt user
                                        img_count = 0
                                        row_count = img_count + 2
                                
                                     else: #for n>1, decrement as usual
                                        img_count += -1
                                        row_count = img_count + 2                               
                                
                                     try: #try to update window...
                                        g.Erase()
                                        graph.DrawImage(filename=f'{ImgArray[img_count]}', location=(0,0))
                                     except IndexError: #..unless we get an indexerror, aka have reached the first image
                                         sg.popup('Reached first image')
                                         pass
                                
                                if MainDrawEvs == 'NextImg':

                                     #reset ClickCount if primed (i.e user has clicked for top-left rectangle pos) when navigating images.
                                     #this prevents drawing bounding box outside image area in edge cases
                                     if ClickCount % 2 == 0: 
                                       ClickCount += 1

                                     PosListArray = [[] for i in range(len(DrawTableDataArray))]
                                     try: #try to update window...
                                        img_count += 1
                                        row_count = img_count + 2
                                        #print(str(img_count) + ': incremented Try state') #diagnostic: print if successul increment
                                        g.Erase()
                                        graph.DrawImage(filename=f'{ImgArray[img_count]}', location=(0,0))
                                
                                     except IndexError: #..unless we get an indexerror, aka have reached the last image
                                         img_count=len(ImgArray)-1
                                         row_count = img_count + 2
                                         print(str(img_count) + ': IndexError reduction') #diagnostic: print if reduced val through IndexError exception
                                         pass

                                if MainDrawEvs =='WriteToExcel':
                                    if LaunchVals['OutputFormat'] == '.xlsx':
                                        Tagbook2.save(filename=dest_filename)
                                    elif LaunchVals['OutputFormat'] == '.json':
                                        Tagbook2.save(filename=dest_filename)
                                        convert_from_file(dest_filename)
                                        os.remove(dest_filename)
                                    elif LaunchVals['OutputFormat'] == '.xlsx + .json':
                                        Tagbook2.save(filename=dest_filename)
                                        convert_from_file(dest_filename)
                                
                                writeVal = [sub[ pathSnip+1:-4] for sub in ImgArray] #remove last 4 chars aka ".png", and first n characters where pathSnip is length of folder location/
                                MainSheet.cell(column=1, row=row_count,  value= writeVal[img_count])
                                MainSheet.cell(row=row_count, column=(button_val+2), value=str(PosListArray[button_val]))

                   if DrawWindDefEvent == 'Back':
                       draw_windowDefine.close()
                       DrawLauncher.un_hide()
                       break

           if DrawLaunchEvent == 'DrawExit': #go back to initial launch screen
               DrawLauncher.close()
               LaunchWindow.UnHide()
               break